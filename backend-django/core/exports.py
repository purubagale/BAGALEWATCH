"""
Excel exports (2026-08-05), ported from v1's 3 "Export to Excel" buttons
in the Backup & Restore modal — exportSiteDetailsXlsx()/
exportSectorDataXlsx()/exportSiteKpiXlsx() (bts_monitor.html
~13222-13292). Column headers and per-row values below match those
functions exactly (site/sector field-for-field), so a v2 export opens
looking like the v1 one an O&M engineer already knows.

v1 hand-rolls its own XLSX writer from scratch (xlsx_download(),
~13060-13218 — builds the raw Open XML parts and zips them with a
custom deflate implementation) because it's a single static HTML file
with no server. v2 has a real Django backend, so this uses openpyxl
instead — same output format, far less code, and gets real column
width/number formatting for free.

**New in v2, not in v1**: v1's 3 export buttons always dumped every site
in the system, no filter. Per explicit user request ("add feature of
selection of all nepal wise, or region or district wise export"), every
export here takes an optional `scope`/`region`/`district` query param —
`scope=all` (default, matches v1's only behavior) or `scope=region`/
`scope=district` to filter to one province or one district before
building the workbook.
"""
import re
from datetime import datetime

import openpyxl
from django.db.models import Count
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Sector, Site

_XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _safe_filename_part(name: str) -> str:
    cleaned = re.sub(r'[^A-Za-z0-9]+', '_', (name or '').strip())
    return cleaned.strip('_') or 'Unnamed'


def _autofit(ws, max_width=42, min_width=8):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        letter = col_cells[0].column_letter
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def _resolve_scope(request):
    """Returns (queryset, label) — label feeds the downloaded filename so
    a scoped export is identifiable at a glance (e.g.
    NT_Site_Details_Bagmati_Province_05-08-2026.xlsx) rather than looking
    identical to a full export."""
    scope = request.query_params.get('scope', 'all')
    qs = Site.objects.all().order_by('id')
    if scope == 'region':
        region = request.query_params.get('region') or ''
        if not region:
            return None, 'region is required when scope=region.'
        return qs.filter(region=region), _safe_filename_part(region)
    if scope == 'district':
        district = request.query_params.get('district') or ''
        if not district:
            return None, 'district is required when scope=district.'
        return qs.filter(district=district), _safe_filename_part(district)
    return qs, 'All_Nepal'


def _build_site_details_workbook(qs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Site Details'
    ws.append(['Site ID', 'Site Name', 'Region', 'City', 'Province / District', 'Latitude', 'Longitude', 'Sector Count'])
    for s in qs.annotate(sector_count=Count('sectors')):
        ws.append([s.id, s.name, s.region, s.city, s.district or s.region, s.lat, s.lng, s.sector_count])
    _autofit(ws)
    return wb


def _build_sector_data_workbook(qs):
    """The 'Lat'/'Long' columns here used to always be the SITE's own
    coordinate, repeated on every one of its sector rows — before Sector
    had any GPS concept of its own, that was the only value there was to
    put there. Now that a sector can carry its own override (2026-08-09,
    "sometimes same sites with multiple sectors may have different lat
    long location as sector expansion" + the same-day import follow-up,
    "when i upload the sector data, also import each sector lat long also
    and store"), this exports `sec.lat`/`sec.lng` when the sector actually
    has an override, falling back to the site's own coordinate only when
    it doesn't — so a re-import of this exact export (see
    site_import.py's `_sector_location_override()`) round-trips a real
    per-sector location correctly instead of flattening it back to the
    site's location on every export/re-import cycle.

    Gained a 'Tech' column 2026-08-09 ("yes for 2g and 3g also need
    sector import") — SECTOR_FIELDS in site_import.py never included
    'tech' at all, so a re-uploaded sector kept whatever Tech it already
    had (or blank, defaulting to 4G in core/sector_expansion.py's
    classification) regardless of what this export showed. Now round-
    trips like every other real column here.

    Gained Carrier/Site Band/Cell Active Status/Site Existence columns
    same day, same follow-up ("need to store all those data also") — see
    Sector.carrier/site_band/cell_active_status/site_existence's
    docstring in models.py. Exported as plain text exactly as stored,
    same round-trip contract as every other column here."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sector Data'
    ws.append([
        'Site ID', 'Cell Name', 'Sector', 'Tech', 'Local Cell ID', 'Lat', 'Long',
        'Height (m)', 'Azimuth (deg)', 'MT (deg)', 'ET (deg)', 'PCI',
        'Carrier', 'Site Band', 'Cell Active Status', 'Site Existence',
    ])
    for s in qs.prefetch_related('sectors'):
        sectors = list(s.sectors.all())
        if not sectors:
            ws.append([s.id, '—', '—', '', None, s.lat, s.lng, None, None, None, None, None, '', '', '', ''])
        else:
            for sec in sectors:
                sec_lat = sec.lat if sec.lat is not None else s.lat
                sec_lng = sec.lng if sec.lng is not None else s.lng
                ws.append([
                    s.id, sec.cell_name or '', sec.sector or '', sec.tech or '', sec.local_cell_id,
                    sec_lat, sec_lng, sec.height, sec.azimuth, sec.mech_tilt, sec.elec_tilt, sec.pci,
                    sec.carrier or '', sec.site_band or '', sec.cell_active_status or '', sec.site_existence or '',
                ])
    _autofit(ws)
    return wb


def _build_site_kpi_workbook(qs):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Sites'
    ws1.append(['Site ID', 'Site Name', 'Region', 'City', 'Latitude', 'Longitude', 'Type', 'Technology', 'Status', 'KPI Status'])
    for s in qs:
        ws1.append([
            s.id, s.name, s.region, s.city, s.lat, s.lng, s.type, s.tech,
            (s.status or '').upper(), 'ENTERED' if s.kpi_entered else 'NOT ENTERED',
        ])
    _autofit(ws1)

    ws2 = wb.create_sheet('KPI Data')
    ws2.append([
        'Site ID', 'Site Name', 'Region', 'Status',
        'RRC Setup SR (%)', 'E-RAB Setup SR (%)', 'Call Setup SR (%)', 'Call Drop Rate (%)', 'Svc Drop Rate (%)',
        'Intra-HO SR (%)', 'Inter-Freq HO SR (%)', 'Inter-RAT HO SR (%)',
        'IP Throughput (Mbps)', 'IP Latency (ms)', 'PRB Utilization (%)', 'Bearer Util (%)', 'License Util (%)', 'Cell Avail (%)',
    ])
    for s in qs:
        no_kpi = not s.kpi_entered
        ws2.append([
            s.id, s.name, s.region, (s.status or '').upper(),
            None if no_kpi else s.rrc, None if no_kpi else s.erab, None if no_kpi else s.call_setup,
            None if no_kpi else s.call_drop, None if no_kpi else s.svc_drop,
            None if no_kpi else s.intra_ho, None if no_kpi else s.inter_ho, None if no_kpi else s.inter_rat,
            None if no_kpi else s.ip_thru, None if no_kpi else s.ip_lat, None if no_kpi else s.prb,
            None if no_kpi else s.bearer_util, None if no_kpi else s.lic_util, None if no_kpi else s.cell_avail,
        ])
    _autofit(ws2)
    return wb


_KIND_BUILDERS = {
    'details': (_build_site_details_workbook, 'NT_Site_Details'),
    'sectors': (_build_sector_data_workbook, 'NT_Site_Sector_Data'),
    'kpi': (_build_site_kpi_workbook, 'NT_Sites_KPI'),
}


class SiteExportXlsxView(APIView):
    """GET /api/v2/export/sites.xlsx?kind=details|sectors|kpi
    &scope=all|region|district&region=<name>&district=<name>

    No extra permission gate beyond being logged in — matches v1 exactly
    (its 3 Excel export buttons never call userCan(), unlike
    exportProject()/restoreProject() in the same modal)."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        kind = request.query_params.get('kind')
        if kind not in _KIND_BUILDERS:
            return Response({'detail': 'kind must be one of: details, sectors, kpi.'}, status=400)

        qs, label = _resolve_scope(request)
        if qs is None:
            return Response({'detail': label}, status=400)
        if not qs.exists():
            return Response({'detail': 'No sites match that scope.'}, status=404)

        builder, prefix = _KIND_BUILDERS[kind]
        wb = builder(qs)
        date_str = datetime.now().strftime('%d-%m-%Y')
        filename = f'{prefix}_{label}_{date_str}.xlsx'

        response = HttpResponse(content_type=_XLSX_MIME)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
